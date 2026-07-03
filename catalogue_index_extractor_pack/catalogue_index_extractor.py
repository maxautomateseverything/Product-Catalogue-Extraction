#!/usr/bin/env python3
"""
Configurable PDF catalogue index extractor.

This version uses row-first, coordinate-based extraction from selectable-text PDFs.
It is designed for catalogue index pages where each product row contains at least:
    - SKU / product code
    - catalogue page reference

Primary engine: pdfplumber word coordinates.
Review workbook: openpyxl, when available.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import textwrap
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import pdfplumber
except ImportError as exc:
    raise SystemExit("Missing dependency: pdfplumber. Install with: py -m pip install pdfplumber") from exc

try:
    import yaml
except ImportError:
    yaml = None

try:
    from PIL import ImageDraw
except Exception:
    ImageDraw = None


# =============================================================================
# Console helpers
# =============================================================================


def log(message: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {message}", flush=True)


def prompt_text(prompt: str, default: Optional[str] = None, required: bool = True) -> str:
    while True:
        suffix = f" [{default}]" if default not in (None, "") else ""
        value = input(f"{prompt}{suffix}: ").strip()
        if not value and default is not None:
            return str(default)
        if value or not required:
            return value
        print("This value is required.")


def prompt_yes_no(prompt: str, default: bool = False) -> bool:
    default_label = "Y/n" if default else "y/N"
    while True:
        value = input(f"{prompt} ({default_label}): ").strip().lower()
        if not value:
            return default
        if value in {"y", "yes"}:
            return True
        if value in {"n", "no"}:
            return False
        print("Please enter y or n.")


def prompt_int(prompt: str, default: Optional[int] = None, minimum: Optional[int] = None) -> int:
    while True:
        suffix = f" [{default}]" if default is not None else ""
        value = input(f"{prompt}{suffix}: ").strip()
        if not value and default is not None:
            return default
        try:
            result = int(value)
            if minimum is not None and result < minimum:
                print(f"Please enter a number >= {minimum}.")
                continue
            return result
        except ValueError:
            print("Please enter a whole number.")


def prompt_float(prompt: str, default: Optional[float] = None, required: bool = True) -> float:
    while True:
        suffix = f" [{default}]" if default is not None else ""
        value = input(f"{prompt}{suffix}: ").strip()
        if not value and default is not None:
            return float(default)
        if not value and not required:
            return 0.0
        try:
            return float(value)
        except ValueError:
            print("Please enter a number, for example: 760")


def prompt_list(prompt: str, minimum: int = 0, help_text: Optional[str] = None) -> List[str]:
    if help_text:
        print(help_text)
    values: List[str] = []
    i = 1
    while True:
        label = "blank to finish" if len(values) >= minimum else "required"
        value = input(f"{prompt} {i} ({label}): ").strip()
        if not value:
            if len(values) >= minimum:
                return values
            print(f"Please provide at least {minimum} value(s).")
            continue
        values.append(value)
        i += 1


# =============================================================================
# Dataclasses
# =============================================================================


@dataclass
class WordLine:
    words: List[Dict[str, Any]]
    top: float
    bottom: float
    x0: float
    x1: float
    text: str


@dataclass
class HeaderOccurrence:
    field_name: str
    header_text: str
    x0: float
    x1: float
    top: float
    bottom: float
    center_x: float
    center_y: float
    pdf_page: int


@dataclass
class ColumnZone:
    field_name: str
    header_text: str
    x0: float
    x1: float
    alignment: str = "left"
    inclusion_mode: str = "anchor_or_overlap"
    value_regex: str = ""
    boundary_warning_threshold: float = 0.75
    header_found: bool = True


@dataclass
class TableBlock:
    block_number: int
    pdf_page: int
    block_x0: float
    block_x1: float
    header_top: float
    header_bottom: float
    data_top: float
    data_bottom: float
    column_zones: Dict[str, ColumnZone]
    detection_mode: str = "auto"
    required_issues: List[str] = field(default_factory=list)
    optional_warnings: List[str] = field(default_factory=list)


# =============================================================================
# Config loading/saving
# =============================================================================


def load_config(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        if path.suffix.lower() in {".yaml", ".yml"}:
            if yaml is None:
                raise SystemExit("PyYAML is required to read YAML config files. Install with: py -m pip install PyYAML")
            return yaml.safe_load(f) or {}
        return json.load(f)


def save_config(config: Dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        if path.suffix.lower() in {".yaml", ".yml"} and yaml is not None:
            yaml.safe_dump(config, f, sort_keys=False, allow_unicode=True)
        else:
            json.dump(config, f, indent=2, ensure_ascii=False)


# =============================================================================
# Utility functions
# =============================================================================


def parse_page_range(page_range: str) -> List[int]:
    pages: List[int] = []
    for part in str(page_range).split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start_s, end_s = part.split("-", 1)
            start = int(start_s.strip())
            end = int(end_s.strip())
            step = 1 if end >= start else -1
            pages.extend(range(start, end + step, step))
        else:
            pages.append(int(part))
    seen = set()
    out = []
    for page in pages:
        if page not in seen:
            seen.add(page)
            out.append(page)
    return out


def collapse_ws(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def normalize_for_match(text: Any, case_sensitive: bool) -> str:
    text = collapse_ws(text)
    return text if case_sensitive else text.lower()


def normalize_sku(text: Any) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(text or "")).upper()


def normalize_generic(text: Any) -> str:
    return collapse_ws(text).upper()


def normalize_page_value(text: Any) -> str:
    raw = collapse_ws(text)
    if not raw:
        return ""
    cleaned = re.sub(r"(?i)\bsee\s+page\b", "", raw).strip()
    cleaned = re.sub(r"(?i)\bpage\b", "", cleaned).strip()
    if re.fullmatch(r"[A-Za-z]?\d+(?:\s*[,/;]\s*[A-Za-z]?\d+)+", cleaned):
        return ";".join(p for p in re.split(r"\s*[,/;]\s*", cleaned) if p)
    return cleaned


def dedupe_preserve(values: Iterable[Any]) -> List[str]:
    out: List[str] = []
    seen = set()
    for value in values:
        value = collapse_ws(value)
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def multiline(values: Iterable[Any]) -> str:
    return "\n".join(dedupe_preserve(values))


def safe_excel_text(value: Any) -> Any:
    if value is None:
        return ""
    s = str(value)
    if s.startswith("="):
        return "'" + s
    return s


def compile_regex(pattern: Optional[str], label: str) -> Optional[re.Pattern[str]]:
    pattern = collapse_ws(pattern)
    if not pattern:
        return None
    try:
        return re.compile(pattern)
    except re.error as exc:
        raise SystemExit(f"Invalid {label} regex: {exc}\nPattern: {pattern}") from exc


def suggest_sku_regex(examples: Sequence[str]) -> str:
    cleaned = [collapse_ws(e).upper() for e in examples if collapse_ws(e)]
    prefixes = []
    has_space = False
    for ex in cleaned:
        if " " in ex:
            has_space = True
        token = re.split(r"[\s\-_/\.]+", ex, maxsplit=1)[0]
        if token:
            prefixes.append(re.escape(token))
    prefixes = dedupe_preserve(prefixes)
    if prefixes and has_space:
        return rf"\b(?:{'|'.join(prefixes)})(?:[\s\-_/\.]+[A-Z0-9]{{1,8}}){{2,5}}\b"
    return r"\b[A-Z0-9][A-Z0-9\-_/\.]{2,40}\b"


def suggest_page_regex(examples: Sequence[str]) -> str:
    return r"(?i)\b(?:see\s+page\s+)?[A-Z]?\d+(?:\s*(?:,|/|;|-)\s*[A-Z]?\d+)*\b"


def word_center(word: Dict[str, Any]) -> Tuple[float, float]:
    return (float(word["x0"]) + float(word["x1"])) / 2, (float(word["top"]) + float(word["bottom"])) / 2


def word_overlaps_zone(word: Dict[str, Any], zone: ColumnZone) -> bool:
    return float(word["x0"]) < zone.x1 and float(word["x1"]) > zone.x0


def word_anchor_in_zone(word: Dict[str, Any], zone: ColumnZone) -> bool:
    x0 = float(word["x0"])
    x1 = float(word["x1"])
    cx = (x0 + x1) / 2
    mode = (zone.inclusion_mode or zone.alignment or "left").lower()
    # Whole words are never split. If a word crosses a boundary, include it and
    # flag it later with boundary_review_for_word. This prevents silent loss when
    # PDF coordinates differ slightly from the visible column alignment.
    overlap = word_overlaps_zone(word, zone)
    if mode in {"left", "word_left"}:
        return zone.x0 <= x0 <= zone.x1 or overlap
    if mode in {"right", "word_right"}:
        return zone.x0 <= x1 <= zone.x1 or overlap
    if mode in {"center", "word_center"}:
        return zone.x0 <= cx <= zone.x1 or overlap
    if mode in {"contained", "word_contained"}:
        return (zone.x0 <= x0 and x1 <= zone.x1) or overlap
    if mode in {"majority", "word_majority"}:
        overlap_width = max(0.0, min(x1, zone.x1) - max(x0, zone.x0))
        width = max(0.0001, x1 - x0)
        return overlap_width / width >= 0.5 or overlap
    return overlap


def boundary_review_for_word(word: Dict[str, Any], zone: ColumnZone) -> bool:
    x0 = float(word["x0"])
    x1 = float(word["x1"])
    if not word_overlaps_zone(word, zone) or (zone.x0 <= x0 and x1 <= zone.x1):
        return False
    overlap = max(0.0, min(x1, zone.x1) - max(x0, zone.x0))
    width = max(0.0001, x1 - x0)
    # Very small coordinate drift can place a visually correct value just outside
    # the header-derived boundary. Only warn when less than 75% of the word is
    # inside the column, which is more likely to be real leakage.
    return (overlap / width) < float(zone.boundary_warning_threshold)


def group_words_to_lines(words: Sequence[Dict[str, Any]], y_tolerance: float = 3.0) -> List[WordLine]:
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: (word_center(w)[1], float(w["x0"])))
    groups: List[List[Dict[str, Any]]] = []
    centers: List[float] = []
    for word in sorted_words:
        _, cy = word_center(word)
        placed = False
        for idx, center in enumerate(centers):
            if abs(cy - center) <= y_tolerance:
                groups[idx].append(word)
                centers[idx] = (centers[idx] * (len(groups[idx]) - 1) + cy) / len(groups[idx])
                placed = True
                break
        if not placed:
            groups.append([word])
            centers.append(cy)

    lines: List[WordLine] = []
    for group in groups:
        sorted_group = sorted(group, key=lambda w: float(w["x0"]))
        text = collapse_ws(" ".join(str(w.get("text", "")) for w in sorted_group))
        if not text:
            continue
        lines.append(
            WordLine(
                words=sorted_group,
                top=min(float(w["top"]) for w in sorted_group),
                bottom=max(float(w["bottom"]) for w in sorted_group),
                x0=min(float(w["x0"]) for w in sorted_group),
                x1=max(float(w["x1"]) for w in sorted_group),
                text=text,
            )
        )
    return sorted(lines, key=lambda ln: (ln.top, ln.x0))


# =============================================================================
# Configuration helpers
# =============================================================================


def get_required_column_config(config: Dict[str, Any], name: str) -> Dict[str, Any]:
    return config.get("required_columns", {}).get(name, {}) or {}


def optional_columns(config: Dict[str, Any]) -> List[Dict[str, Any]]:
    return config.get("optional_columns", []) or []


def all_column_configs(config: Dict[str, Any]) -> List[Tuple[str, Dict[str, Any]]]:
    cols = [
        ("sku", get_required_column_config(config, "sku")),
        ("page", get_required_column_config(config, "page")),
    ]
    for col in optional_columns(config):
        cols.append((col["output_name"], col))
    return cols


def field_alignment(config: Dict[str, Any], field_name: str) -> str:
    for name, col in all_column_configs(config):
        if name == field_name:
            return str(col.get("alignment", "left")).lower()
    return "left"


def field_inclusion_mode(config: Dict[str, Any], field_name: str) -> str:
    for name, col in all_column_configs(config):
        if name == field_name:
            return str(col.get("inclusion_mode", col.get("alignment", "left"))).lower()
    return "left"


def field_header_text(config: Dict[str, Any], field_name: str) -> str:
    if field_name in {"sku", "page"}:
        return str(get_required_column_config(config, field_name).get("header_text", field_name))
    for col in optional_columns(config):
        if col.get("output_name") == field_name:
            return str(col.get("header_text", field_name))
    return field_name


def has_center_aligned_columns(config: Dict[str, Any]) -> bool:
    return any(field_alignment(config, name) == "center" for name, _ in all_column_configs(config))


# =============================================================================
# Header and block detection
# =============================================================================


def find_exact_header_occurrences(
    words: Sequence[Dict[str, Any]],
    header_text: str,
    field_name: str,
    pdf_page: int,
    case_sensitive: bool,
    y_tolerance: float,
) -> List[HeaderOccurrence]:
    tokens = [normalize_for_match(t, case_sensitive) for t in collapse_ws(header_text).split()]
    if not tokens:
        return []
    found: List[HeaderOccurrence] = []
    for line in group_words_to_lines(words, y_tolerance):
        line_words = sorted(line.words, key=lambda w: float(w["x0"]))
        line_tokens = [normalize_for_match(w.get("text", ""), case_sensitive) for w in line_words]
        for i in range(0, len(line_tokens) - len(tokens) + 1):
            if line_tokens[i : i + len(tokens)] == tokens:
                matched = line_words[i : i + len(tokens)]
                x0 = min(float(w["x0"]) for w in matched)
                x1 = max(float(w["x1"]) for w in matched)
                top = min(float(w["top"]) for w in matched)
                bottom = max(float(w["bottom"]) for w in matched)
                found.append(
                    HeaderOccurrence(
                        field_name=field_name,
                        header_text=header_text,
                        x0=x0,
                        x1=x1,
                        top=top,
                        bottom=bottom,
                        center_x=(x0 + x1) / 2,
                        center_y=(top + bottom) / 2,
                        pdf_page=pdf_page,
                    )
                )
    return sorted(found, key=lambda h: (h.center_y, h.center_x))


def is_same_header_line(a: HeaderOccurrence, b: HeaderOccurrence, tolerance: float) -> bool:
    return abs(a.center_y - b.center_y) <= tolerance


def boundary_between(left_h: HeaderOccurrence, left_align: str, right_h: HeaderOccurrence, right_align: str) -> float:
    # User-confirmed auto-boundary logic.
    # If both columns are right aligned, the boundary is x1 of the left header.
    # Otherwise the boundary is x0 of the right header.
    if left_align == "right" and right_align == "right":
        return left_h.x1
    return right_h.x0


def build_column_zones_from_headers(
    block_x0: float,
    block_x1: float,
    headers_by_field: Dict[str, HeaderOccurrence],
    config: Dict[str, Any],
) -> Dict[str, ColumnZone]:
    sorted_items = sorted(headers_by_field.items(), key=lambda item: item[1].x0)
    zones: Dict[str, ColumnZone] = {}
    left_boundaries: Dict[str, float] = {}
    right_boundaries: Dict[str, float] = {}

    for idx, (field, header) in enumerate(sorted_items):
        if idx == 0:
            left_boundaries[field] = header.x0
        else:
            prev_field, prev_h = sorted_items[idx - 1]
            b = boundary_between(prev_h, field_alignment(config, prev_field), header, field_alignment(config, field))
            right_boundaries[prev_field] = b
            left_boundaries[field] = b if (field_alignment(config, prev_field) == "right" and field_alignment(config, field) == "right") else header.x0

    last_field, _last_header = sorted_items[-1]
    right_boundaries[last_field] = block_x1

    for field, header in sorted_items:
        zones[field] = ColumnZone(
            field_name=field,
            header_text=header.header_text,
            x0=float(left_boundaries[field]),
            x1=float(right_boundaries[field]),
            alignment=field_alignment(config, field),
            inclusion_mode=field_inclusion_mode(config, field),
            value_regex=str(next((c.get("value_regex", "") for name, c in all_column_configs(config) if name == field), "") or ""),
            boundary_warning_threshold=float(config.get("advanced", {}).get("boundary_overlap_warning_threshold", 0.75)),
            header_found=True,
        )
    return zones


def detect_auto_blocks(page_width: float, pdf_page: int, words: Sequence[Dict[str, Any]], page_height: float, config: Dict[str, Any]) -> Tuple[List[TableBlock], List[Dict[str, Any]]]:
    advanced = config.get("advanced", {}) or {}
    case_sensitive = bool(config.get("header_matching", {}).get("case_sensitive", False))
    header_y_tol = float(advanced.get("header_y_tolerance", 6.0))
    data_start_padding = float(advanced.get("data_start_padding", 1.0))
    data_bottom = float(advanced["data_bottom"])

    header_records: List[Dict[str, Any]] = []
    header_occurrences_by_field: Dict[str, List[HeaderOccurrence]] = {}
    for field, col_cfg in all_column_configs(config):
        header = str(col_cfg.get("header_text", field))
        occ = find_exact_header_occurrences(words, header, field, pdf_page, case_sensitive, y_tolerance=3.0)
        header_occurrences_by_field[field] = occ
        for h in occ:
            header_records.append({
                "source_pdf_page": pdf_page,
                "field_name": field,
                "header_text": h.header_text,
                "x0": round(h.x0, 3),
                "x1": round(h.x1, 3),
                "top": round(h.top, 3),
                "bottom": round(h.bottom, 3),
                "center_x": round(h.center_x, 3),
                "center_y": round(h.center_y, 3),
                "detection_mode": "auto_header_search",
            })

    sku_headers = header_occurrences_by_field.get("sku", [])
    page_headers = header_occurrences_by_field.get("page", [])
    pairs: List[Tuple[HeaderOccurrence, HeaderOccurrence]] = []
    used_page_ids = set()
    for sku_h in sorted(sku_headers, key=lambda h: h.x0):
        candidates = [
            p for p in page_headers
            if id(p) not in used_page_ids and p.x0 > sku_h.x0 and is_same_header_line(sku_h, p, header_y_tol)
        ]
        if not candidates:
            continue
        page_h = sorted(candidates, key=lambda p: p.x0 - sku_h.x0)[0]
        used_page_ids.add(id(page_h))
        pairs.append((sku_h, page_h))
    pairs = sorted(pairs, key=lambda p: p[0].x0)

    blocks: List[TableBlock] = []
    for idx, (sku_h, page_h) in enumerate(pairs, start=1):
        prev_pair = pairs[idx - 2] if idx > 1 else None
        next_pair = pairs[idx] if idx < len(pairs) else None
        block_x0 = 0.0 if prev_pair is None else (prev_pair[1].x1 + sku_h.x0) / 2
        block_x1 = page_width if next_pair is None else (page_h.x1 + next_pair[0].x0) / 2
        headers = {"sku": sku_h, "page": page_h}

        # Add optional headers that sit in this block and on the same header row.
        for col in optional_columns(config):
            field = col["output_name"]
            cands = [
                h for h in header_occurrences_by_field.get(field, [])
                if block_x0 <= h.center_x <= block_x1 and is_same_header_line(sku_h, h, header_y_tol)
            ]
            if cands:
                headers[field] = sorted(cands, key=lambda h: h.x0)[0]

        zones = build_column_zones_from_headers(block_x0, block_x1, headers, config)
        warnings = []
        for col in optional_columns(config):
            if col["output_name"] not in zones:
                warnings.append(f"optional_header_not_found:{col['output_name']}")

        blocks.append(TableBlock(
            block_number=idx,
            pdf_page=pdf_page,
            block_x0=block_x0,
            block_x1=block_x1,
            header_top=min(sku_h.top, page_h.top),
            header_bottom=max(sku_h.bottom, page_h.bottom),
            data_top=max(sku_h.bottom, page_h.bottom) + data_start_padding,
            data_bottom=data_bottom,
            column_zones=zones,
            detection_mode="auto",
            optional_warnings=warnings,
        ))

    return blocks, header_records


def manual_blocks_from_config(page_width: float, page_height: float, pdf_page: int, config: Dict[str, Any]) -> List[TableBlock]:
    blocks: List[TableBlock] = []
    data_bottom = float(config.get("advanced", {}).get("data_bottom", page_height))
    for idx, block_cfg in enumerate(config.get("manual_coordinate_blocks", []) or [], start=1):
        zones: Dict[str, ColumnZone] = {}
        for field_name, z in (block_cfg.get("columns", {}) or {}).items():
            zones[field_name] = ColumnZone(
                field_name=field_name,
                header_text=z.get("header_text", field_header_text(config, field_name)),
                x0=float(z["x0"]),
                x1=float(z["x1"]),
                alignment=str(z.get("alignment", field_alignment(config, field_name))).lower(),
                inclusion_mode=str(z.get("inclusion_mode", z.get("alignment", field_alignment(config, field_name)))).lower(),
                value_regex=str(z.get("value_regex", "") or ""),
                boundary_warning_threshold=float(config.get("advanced", {}).get("boundary_overlap_warning_threshold", 0.75)),
                header_found=False,
            )
        all_xs = [x for z in zones.values() for x in (z.x0, z.x1)] or [0.0, page_width]
        blocks.append(TableBlock(
            block_number=idx,
            pdf_page=pdf_page,
            block_x0=float(block_cfg.get("block_x0", min(all_xs))),
            block_x1=float(block_cfg.get("block_x1", max(all_xs))),
            header_top=float(block_cfg.get("data_top", 0.0)),
            header_bottom=float(block_cfg.get("data_top", 0.0)),
            data_top=float(block_cfg.get("data_top", 0.0)),
            data_bottom=float(block_cfg.get("data_bottom", data_bottom)),
            column_zones=zones,
            detection_mode="manual",
        ))
    return blocks


# =============================================================================
# Row-first extraction
# =============================================================================


def line_words_in_block(line: WordLine, block: TableBlock) -> List[Dict[str, Any]]:
    out = []
    for w in line.words:
        cx, cy = word_center(w)
        if block.block_x0 <= cx <= block.block_x1 and block.data_top <= cy <= block.data_bottom:
            out.append(w)
    return out


def text_from_zone_on_line(line: WordLine, zone: ColumnZone) -> Tuple[str, List[str], List[Dict[str, Any]]]:
    selected: List[Dict[str, Any]] = []
    warnings: List[str] = []
    for word in sorted(line.words, key=lambda w: float(w["x0"])):
        if word_anchor_in_zone(word, zone):
            selected.append(word)
            if boundary_review_for_word(word, zone):
                warnings.append(f"boundary_overlap:{zone.field_name}:{word.get('text', '')}")
    text = collapse_ws(" ".join(str(w.get("text", "")) for w in selected))
    return text, warnings, selected


def build_stop_regexes(config: Dict[str, Any], page_regex: Optional[re.Pattern[str]]) -> List[Tuple[re.Pattern[str], str]]:
    stops: List[Tuple[re.Pattern[str], str]] = []
    # Do not use the page regex as a SKU stop pattern because numeric tokens are
    # often valid inside SKUs, e.g. GW 20 923. Optional-column value regexes are
    # safer stop signals for neighbouring-column leakage, such as 100/5200.
    for col in optional_columns(config):
        pat = collapse_ws(col.get("value_regex", ""))
        # Use only columns that are safe stop signals for SKU leakage. Numeric-only
        # patterns such as pallet quantities are unsafe because SKU tokens often
        # contain numbers. By default, slash-containing patterns such as Pack/carton
        # are used as stop signals. This can be overridden with use_as_sku_stop.
        use_as_stop = col.get("use_as_sku_stop")
        if use_as_stop is None:
            use_as_stop = "/" in pat
        if pat and bool(use_as_stop):
            compiled = compile_regex(pat, f"optional column {col.get('output_name')} value")
            if compiled is not None:
                must_contain = str(col.get("sku_stop_token_must_contain", "/" if "/" in pat else ""))
                stops.append((compiled, must_contain))
    return stops


def truncate_text_at_stop_pattern(text: str, stop_regexes: Sequence[Tuple[re.Pattern[str], str]]) -> str:
    tokens = text.split()
    kept: List[str] = []
    for token in tokens:
        stop_now = False
        for rx, must_contain in stop_regexes:
            if must_contain and must_contain not in token:
                continue
            if rx.fullmatch(token) or rx.search(token):
                stop_now = True
                break
        if stop_now:
            break
        kept.append(token)
    return collapse_ws(" ".join(kept))


def clean_sku_text_for_matching(text: str, config: Dict[str, Any]) -> str:
    rules = config.get("sku_rules", {}) or {}
    out = str(text or "")
    if bool(rules.get("uppercase_only", False)):
        # Remove lowercase letters after candidate evidence has been captured.
        out = re.sub(r"[a-z]", "", out)
    allowed = str(rules.get("allowed_characters", "A-Z0-9 space hyphen slash dot underscore plus") or "")
    # Keep common configured punctuation. Unusual symbols like circled letters are removed.
    keep_chars = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789")
    if "space" in allowed.lower():
        keep_chars.add(" ")
    if "hyphen" in allowed.lower() or "-" in allowed:
        keep_chars.add("-")
    if "slash" in allowed.lower() or "/" in allowed:
        keep_chars.add("/")
    if "dot" in allowed.lower() or "." in allowed:
        keep_chars.add(".")
    if "underscore" in allowed.lower() or "_" in allowed:
        keep_chars.add("_")
    if "plus" in allowed.lower() or "+" in allowed:
        keep_chars.add("+")
    out = "".join(ch if ch.upper() in keep_chars or ch in keep_chars else " " for ch in out.upper())
    return collapse_ws(out)


def extract_sku_from_text(raw_text: str, config: Dict[str, Any], sku_regex: Optional[re.Pattern[str]], stop_regexes: Sequence[Tuple[re.Pattern[str], str]]) -> Tuple[str, str, str, str]:
    """Return (sku_clean, sku_normalized, raw_evidence, status_warning)."""
    raw_evidence = collapse_ws(raw_text)
    if not raw_evidence:
        return "", "", "", ""
    text_for_matching = clean_sku_text_for_matching(raw_evidence, config)
    text_for_matching = truncate_text_at_stop_pattern(text_for_matching, stop_regexes)
    if not text_for_matching:
        return "", "", raw_evidence, ""
    if sku_regex is None:
        candidate = text_for_matching
    else:
        matches = list(sku_regex.finditer(text_for_matching))
        if not matches:
            return "", "", raw_evidence, "sku_regex_no_match"
        candidate = collapse_ws(matches[0].group(0))
    return candidate, normalize_sku(candidate), raw_evidence, "" if candidate else "sku_empty_after_cleaning"


def extract_page_from_text(raw_text: str, page_regex: Optional[re.Pattern[str]]) -> Tuple[str, str, str]:
    text = collapse_ws(raw_text)
    if not text:
        return "", "", ""
    if page_regex is None:
        page = text
        status = ""
    else:
        m = page_regex.search(text)
        if not m:
            return text, normalize_page_value(text), "page_value_did_not_match_regex"
        page = collapse_ws(m.group(0))
        status = ""
    return page, normalize_page_value(page), status


def validate_optional_value(raw_text: str, col: Dict[str, Any]) -> Tuple[str, str, str]:
    raw = collapse_ws(raw_text)
    if not raw:
        return "", "", "missing"
    pat = collapse_ws(col.get("value_regex", ""))
    if not pat:
        return raw, raw, "confirmed"
    rx = compile_regex(pat, f"optional column {col.get('output_name')} value")
    assert rx is not None
    if rx.fullmatch(raw):
        return raw, raw, "confirmed"
    # If a boundary overlap brings in neighbouring text, try extracting the first
    # matching value from inside the raw string. Anchored user regexes are safely
    # relaxed for this fallback only.
    relaxed_pattern = pat.strip()
    if relaxed_pattern.startswith("^"):
        relaxed_pattern = relaxed_pattern[1:]
    if relaxed_pattern.endswith("$"):
        relaxed_pattern = relaxed_pattern[:-1]
    try:
        relaxed_rx = re.compile(relaxed_pattern)
        m = relaxed_rx.search(raw)
        if m:
            return raw, collapse_ws(m.group(0)), "confirmed_from_raw"
    except re.error:
        pass
    action = str(col.get("invalid_value_action", "blank_and_warn"))
    if action == "keep_and_warn":
        return raw, raw, "invalid_kept"
    return raw, "", "invalid_blank"


def row_matches_ignore_patterns(line_text: str, config: Dict[str, Any]) -> bool:
    patterns = config.get("ignore_row_patterns", []) or []
    text = collapse_ws(line_text)
    for pat in patterns:
        rx = compile_regex(str(pat), "ignore row")
        if rx and rx.search(text):
            return True
    return False


def extract_rows_from_block(
    pdf_page: int,
    words: Sequence[Dict[str, Any]],
    block: TableBlock,
    config: Dict[str, Any],
    sku_regex: Optional[re.Pattern[str]],
    page_regex: Optional[re.Pattern[str]],
) -> Tuple[List[Dict[str, Any]], List[str], List[str], List[Dict[str, Any]]]:
    advanced = config.get("advanced", {}) or {}
    y_tol = float(advanced.get("line_y_tolerance", 3.0))
    stop_regexes = build_stop_regexes(config, page_regex)
    required_issues = list(block.required_issues)
    optional_warnings = list(block.optional_warnings)
    raw_audit_rows: List[Dict[str, Any]] = []
    rows: List[Dict[str, Any]] = []

    if "sku" not in block.column_zones or "page" not in block.column_zones:
        required_issues.append("required_column_zone_missing")
        return rows, required_issues, optional_warnings, raw_audit_rows

    block_words = [w for w in words if block.block_x0 <= word_center(w)[0] <= block.block_x1 and block.data_top <= word_center(w)[1] <= block.data_bottom]
    for row_number, line in enumerate(group_words_to_lines(block_words, y_tol), start=1):
        if row_matches_ignore_patterns(line.text, config):
            continue

        sku_text, sku_boundary_warnings, _sku_words = text_from_zone_on_line(line, block.column_zones["sku"])
        page_text, page_boundary_warnings, _page_words = text_from_zone_on_line(line, block.column_zones["page"])
        sku, sku_norm, sku_raw, sku_status = extract_sku_from_text(sku_text, config, sku_regex, stop_regexes)
        page_original, page_normalized, page_status = extract_page_from_text(page_text, page_regex)

        # Raw SKU-column audit is restricted to the detected SKU column only.
        if sku_raw:
            raw_audit_rows.append({
                "source_pdf_page": pdf_page,
                "source_table_block": block.block_number,
                "row_number_in_block": row_number,
                "sku_column_text": sku_raw,
                "sku_candidate": sku,
                "sku_normalized": sku_norm,
                "in_structured_output_before_raw_add": bool(sku and page_original),
                "added_to_registry_from_raw_text": False,
            })

        # Footer page numbers and title rows with no SKU are ignored as product rows.
        if not sku and not page_original:
            continue
        if not sku and page_original:
            continue

        row_required_issues: List[str] = []
        row_optional_warnings: List[str] = []
        if sku_status:
            row_required_issues.append(sku_status)
        if sku and not page_original:
            row_required_issues.append("missing_catalogue_page_for_sku")
        if page_status:
            row_required_issues.append(page_status)
        for warn in sku_boundary_warnings + page_boundary_warnings:
            row_required_issues.append(warn)

        row: Dict[str, Any] = {
            "sku_raw": sku_raw,
            "sku": sku,
            "sku_normalized": sku_norm,
            "catalogue_page_original": page_original,
            "catalogue_page_normalized": page_normalized,
            "source_pdf_page": pdf_page,
            "source_catalogue_page": make_source_catalogue_page(pdf_page, config),
            "source_table_block": block.block_number,
            "row_number_in_block": row_number,
            "source_method": "pdfplumber_row_coordinate",
            "detection_mode": block.detection_mode,
            "row_top": round(line.top, 3),
            "row_bottom": round(line.bottom, 3),
            "source_row_text": line.text,
            "source_sku_column_text": sku_text,
            "source_page_column_text": page_text,
        }

        for col in optional_columns(config):
            name = col["output_name"]
            raw_value = ""
            clean_value = ""
            status = "missing"
            if name in block.column_zones:
                opt_text, opt_boundary_warnings, _ = text_from_zone_on_line(line, block.column_zones[name])
                raw_value, clean_value, status = validate_optional_value(opt_text, col)
                for warn in opt_boundary_warnings:
                    row_optional_warnings.append(warn)
                if status.startswith("invalid"):
                    row_optional_warnings.append(f"optional_invalid:{name}:{raw_value}")
            row[f"{name}_raw"] = raw_value
            row[name] = clean_value
            row[f"{name}_status"] = status

        row["confidence_status"] = "needs_review" if row_required_issues else "confirmed"
        row["required_issues"] = multiline(row_required_issues)
        row["optional_warnings"] = multiline(row_optional_warnings)
        row["review_reason"] = multiline(row_required_issues + row_optional_warnings)
        if row_required_issues:
            required_issues.append(f"block_{block.block_number}:row_{row_number}:" + ";".join(row_required_issues))
        if row_optional_warnings:
            optional_warnings.append(f"block_{block.block_number}:row_{row_number}:" + ";".join(row_optional_warnings))
        rows.append(row)

    return rows, required_issues, optional_warnings, raw_audit_rows


# =============================================================================
# Raw audit and validation groups
# =============================================================================


def add_raw_column_candidates(index_rows: List[Dict[str, Any]], raw_audit_rows: List[Dict[str, Any]], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    structured_norms = {r.get("sku_normalized", "") for r in index_rows if r.get("sku_normalized") and r.get("catalogue_page_original")}
    unresolved: List[Dict[str, Any]] = []
    for audit in raw_audit_rows:
        norm = audit.get("sku_normalized", "")
        audit["in_final_structured_rows"] = norm in structured_norms
        if norm and norm not in structured_norms:
            audit["added_to_registry_from_raw_text"] = True
            raw_row = {
                "sku_raw": audit.get("sku_column_text", ""),
                "sku": audit.get("sku_candidate", ""),
                "sku_normalized": norm,
                "catalogue_page_original": "",
                "catalogue_page_normalized": "",
                "source_pdf_page": audit.get("source_pdf_page", ""),
                "source_catalogue_page": make_source_catalogue_page(int(audit.get("source_pdf_page", 0) or 0), config),
                "source_table_block": audit.get("source_table_block", ""),
                "row_number_in_block": audit.get("row_number_in_block", ""),
                "source_method": "sku_column_raw_audit",
                "detection_mode": "raw_sku_column_only",
                "row_top": "",
                "row_bottom": "",
                "source_row_text": audit.get("sku_column_text", ""),
                "source_sku_column_text": audit.get("sku_column_text", ""),
                "source_page_column_text": "",
                "confidence_status": "needs_review",
                "required_issues": "raw_sku_column_candidate_not_in_structured_rows",
                "optional_warnings": "",
                "review_reason": "raw_sku_column_candidate_not_in_structured_rows",
            }
            for col in optional_columns(config):
                name = col["output_name"]
                raw_row[f"{name}_raw"] = ""
                raw_row[name] = ""
                raw_row[f"{name}_status"] = "missing"
            index_rows.append(raw_row)
            unresolved.append(raw_row)
            structured_norms.add(norm)
        else:
            audit["added_to_registry_from_raw_text"] = False
    return unresolved


def build_sku_registry(index_rows: List[Dict[str, Any]], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in index_rows:
        norm = row.get("sku_normalized", "")
        if norm:
            grouped[norm].append(row)

    registry: List[Dict[str, Any]] = []
    for norm, rows in sorted(grouped.items(), key=lambda item: item[0]):
        first = next((r for r in rows if r.get("sku")), rows[0])
        required_issues = dedupe_preserve(part for r in rows for part in str(r.get("required_issues", "")).split("\n") if part)
        optional_warns = dedupe_preserve(part for r in rows for part in str(r.get("optional_warnings", "")).split("\n") if part)
        reg: Dict[str, Any] = {
            "sku": first.get("sku", norm),
            "sku_raw_examples": ";".join(dedupe_preserve(r.get("sku_raw", "") for r in rows)),
            "sku_normalized": norm,
            "catalogue_pages_original": ";".join(dedupe_preserve(r.get("catalogue_page_original", "") for r in rows)),
            "catalogue_pages_normalized": ";".join(dedupe_preserve(r.get("catalogue_page_normalized", "") for r in rows)),
            "source_pdf_pages": ";".join(dedupe_preserve(str(r.get("source_pdf_page", "")) for r in rows)),
            "source_catalogue_pages": ";".join(dedupe_preserve(str(r.get("source_catalogue_page", "")) for r in rows)),
            "source_table_blocks": ";".join(dedupe_preserve(str(r.get("source_table_block", "")) for r in rows)),
            "source_methods": ";".join(dedupe_preserve(r.get("source_method", "") for r in rows)),
            "confidence_status": "needs_review" if required_issues else "confirmed",
            "required_issues": multiline(required_issues),
            "optional_warnings": multiline(optional_warns),
            "occurrence_count": len(rows),
        }
        for col in optional_columns(config):
            name = col["output_name"]
            reg[f"{name}_raw"] = ";".join(dedupe_preserve(r.get(f"{name}_raw", "") for r in rows))
            reg[name] = ";".join(dedupe_preserve(r.get(name, "") for r in rows))
            statuses = dedupe_preserve(r.get(f"{name}_status", "") for r in rows)
            reg[f"{name}_statuses"] = ";".join(statuses)
        registry.append(reg)
    return registry


def validate_groups(index_rows: List[Dict[str, Any]], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    groups = config.get("example_validation_groups", []) or []
    results: List[Dict[str, Any]] = []
    for idx, group in enumerate(groups, start=1):
        code = group.get("code", group.get("sku", ""))
        code_norm = normalize_sku(code)
        supplied = {k: v for k, v in group.items() if k not in {"code", "sku"}}
        candidates = [r for r in index_rows if r.get("sku_normalized") == code_norm]
        matched = None
        mismatch_details: List[str] = []
        for row in candidates:
            checks = []
            for key, expected in supplied.items():
                field = "catalogue_page_original" if key in {"page", "catalogue_page"} else key
                actual = row.get(field, "")
                if normalize_generic(actual) == normalize_generic(expected) or normalize_page_value(actual) == normalize_page_value(expected):
                    checks.append(True)
                else:
                    checks.append(False)
                    mismatch_details.append(f"{field}:expected={expected};actual={actual}")
            if all(checks):
                matched = row
                break
        status = "found" if matched else ("sku_not_found" if not candidates else "row_group_not_found")
        results.append({
            "validation_group_number": idx,
            "code": code,
            "sku_normalized": code_norm,
            "status": status,
            "expected_fields": json.dumps(supplied, ensure_ascii=False),
            "matched_source_pdf_page": matched.get("source_pdf_page", "") if matched else "",
            "matched_source_table_block": matched.get("source_table_block", "") if matched else "",
            "matched_row_number_in_block": matched.get("row_number_in_block", "") if matched else "",
            "review_reason": "" if matched else multiline(dedupe_preserve(mismatch_details) or [status]),
        })
    return results


# =============================================================================
# Interactive setup
# =============================================================================


def print_first_candidate_coordinates(pdf_path: Path, first_pdf_page: int, sku_regex: Optional[re.Pattern[str]], page_regex: Optional[re.Pattern[str]]) -> None:
    log(f"Inspecting PDF page {first_pdf_page} for example row coordinates")
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            page = pdf.pages[first_pdf_page - 1]
            words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False, use_text_flow=False)
            lines = group_words_to_lines(words, 3.0)
            printed = 0
            print("\nExample visual rows / candidate coordinates:")
            for line in lines:
                sku_hit = bool(sku_regex and sku_regex.search(clean_sku_text_for_matching(line.text, {"sku_rules": {"uppercase_only": True}})))
                page_hit = bool(page_regex and page_regex.search(line.text))
                if sku_hit or page_hit:
                    print(f"  text='{line.text[:140]}'")
                    print(f"    x0={line.x0:.2f} x1={line.x1:.2f} top={line.top:.2f} bottom={line.bottom:.2f}")
                    printed += 1
                    if printed >= 12:
                        break
            print("")
    except Exception as exc:
        print(f"Could not inspect candidate coordinates: {exc}")


def prompt_float_range(label: str) -> Tuple[float, float]:
    while True:
        value = input(f"{label} x-range as x0,x1: ").strip()
        try:
            x0_s, x1_s = value.split(",", 1)
            x0 = float(x0_s.strip())
            x1 = float(x1_s.strip())
            if x1 <= x0:
                print("x1 must be greater than x0.")
                continue
            return x0, x1
        except Exception:
            print("Please enter two numbers separated by a comma, for example: 35,90")


def prompt_manual_blocks(config: Dict[str, Any]) -> None:
    pages = parse_page_range(config["index_pdf_pages"])
    sku_rx = compile_regex(config.get("sku_detection", {}).get("sku_regex", ""), "SKU")
    page_rx = compile_regex(config.get("page_detection", {}).get("page_regex", ""), "page")
    print_first_candidate_coordinates(Path(config["input_pdf"]), pages[0], sku_rx, page_rx)
    expected = int(config.get("expected_table_blocks_per_page", 1))
    count = prompt_int("How many manual table blocks do you want to define?", expected, minimum=1)
    blocks = []
    for i in range(1, count + 1):
        print(f"\nManual coordinates for table block {i}")
        columns: Dict[str, Any] = {}
        for field, col_cfg in all_column_configs(config):
            x0, x1 = prompt_float_range(f"{field} column")
            columns[field] = {
                "x0": x0,
                "x1": x1,
                "header_text": col_cfg.get("header_text", field),
                "alignment": col_cfg.get("alignment", "left"),
                "inclusion_mode": col_cfg.get("inclusion_mode", col_cfg.get("alignment", "left")),
                "value_regex": col_cfg.get("value_regex", ""),
            }
        data_top = prompt_float("Data top y-coordinate", default=0.0)
        data_bottom = prompt_float("Data bottom y-coordinate", default=float(config.get("advanced", {}).get("data_bottom", 0) or 0))
        xs = [v for z in columns.values() for v in (float(z["x0"]), float(z["x1"]))]
        blocks.append({"block_x0": min(xs), "block_x1": max(xs), "data_top": data_top, "data_bottom": data_bottom, "columns": columns})
    config["manual_coordinate_blocks"] = blocks
    config["extraction_mode"] = "manual"


def build_interactive_config() -> Dict[str, Any]:
    print("\nCatalogue Index Extractor - Interactive Setup\n")
    input_pdf = prompt_text("Input PDF path")
    output_folder = prompt_text("Output folder", default=str(Path(input_pdf).with_name("index_output")))
    pages = prompt_text("Index PDF page range", default="1293-1364")
    case_sensitive = prompt_yes_no("Should header matching be case-sensitive?", default=False)
    sku_header = prompt_text("Exact product-code/SKU column header text", default="Code")
    page_header = prompt_text("Exact catalogue-page column header text", default="Page")
    expected_blocks = prompt_int("Expected product-code/page table blocks per page", default=1, minimum=1)
    data_bottom = prompt_float("Manual data bottom y-coordinate / table cutoff", required=True)

    def column_alignment_prompt(field: str, default: str = "left") -> str:
        while True:
            val = prompt_text(f"Alignment for {field} column: left/right/center", default=default).lower()
            if val in {"left", "right", "center"}:
                return val
            print("Please enter left, right, or center.")

    sku_alignment = column_alignment_prompt("SKU", "left")
    page_alignment = column_alignment_prompt("Page", "left")

    optional_cols = []
    if prompt_yes_no("Do you want to configure optional columns?", default=True):
        while True:
            header = prompt_text("Optional column exact header text", required=False)
            if not header:
                break
            suggested = re.sub(r"[^A-Za-z0-9]+", "_", header).strip("_").lower() or "optional_column"
            name = prompt_text("Output column name", default=suggested)
            alignment = column_alignment_prompt(name, "left")
            value_regex = prompt_text("Optional value regex (blank = no regex)", required=False)
            optional_cols.append({
                "output_name": name,
                "header_text": header,
                "alignment": alignment,
                "inclusion_mode": alignment,
                "value_regex": value_regex,
                "invalid_value_action": "blank_and_warn",
            })

    sku_examples = prompt_list("Positive product-code example", minimum=3)
    negative_examples = prompt_list("Optional negative/non-product row example", minimum=0)
    suggested_sku = suggest_sku_regex(sku_examples)
    print("\nSuggested SKU regex:")
    print(suggested_sku)
    sku_regex = suggested_sku if prompt_yes_no("Use suggested SKU regex?", default=True) else prompt_text("Enter SKU regex")
    compile_regex(sku_regex, "SKU")

    page_examples = prompt_list("Catalogue page-value example", minimum=1)
    suggested_page = suggest_page_regex(page_examples)
    print("\nSuggested page-value regex:")
    print(suggested_page)
    page_regex = suggested_page if prompt_yes_no("Use suggested page regex?", default=True) else prompt_text("Enter page regex, or blank to keep any non-empty page-column text", required=False)
    compile_regex(page_regex, "page")

    ignore_patterns = []
    if negative_examples:
        for ex in negative_examples:
            ignore_patterns.append("^" + re.escape(ex) + "$" )

    groups = []
    if prompt_yes_no("Do you want to enter validation groups?", default=True):
        print("Enter groups as JSON objects. Example: {\"code\":\"DX 10 016 R\",\"pack_carton\":\"100/6400\",\"page\":\"344\"}")
        while True:
            raw = prompt_text("Validation group JSON", required=False)
            if not raw:
                break
            try:
                groups.append(json.loads(raw))
            except json.JSONDecodeError as exc:
                print(f"Invalid JSON: {exc}")

    config: Dict[str, Any] = {
        "input_pdf": input_pdf,
        "output_folder": output_folder,
        "index_pdf_pages": pages,
        "page_source_of_truth": "pdf_page_number",
        "required_columns": {
            "sku": {"header_text": sku_header, "alignment": sku_alignment, "inclusion_mode": sku_alignment},
            "page": {"header_text": page_header, "alignment": page_alignment, "inclusion_mode": page_alignment},
        },
        "optional_columns": optional_cols,
        "expected_table_blocks_per_page": expected_blocks,
        "header_matching": {"case_sensitive": case_sensitive},
        "sku_detection": {"positive_examples": sku_examples, "negative_examples": negative_examples, "sku_regex": sku_regex},
        "sku_rules": {
            "uppercase_only": True,
            "allowed_characters": "A-Z0-9 space hyphen slash dot underscore plus",
        },
        "page_detection": {"positive_examples": page_examples, "page_regex": page_regex, "keep_original_and_normalized": True},
        "ignore_row_patterns": ignore_patterns,
        "example_validation_groups": groups,
        "extraction_mode": "auto",
        "debug_images": {"enabled": True, "only_issue_pages": True, "label_zones": True},
        "review_files": {"excel_safe": True, "write_separate_csvs": False},
        "advanced": {"header_y_tolerance": 6.0, "line_y_tolerance": 3.0, "data_start_padding": 1.0, "x_tolerance": 1, "y_tolerance": 3, "data_bottom": data_bottom},
    }

    if has_center_aligned_columns(config) or prompt_yes_no("Use manual coordinate mode?", default=False):
        print("Manual coordinate mode is required/selected. Centre-aligned columns cannot be auto-zoned safely.")
        prompt_manual_blocks(config)

    if prompt_yes_no("Save this config file?", default=True):
        save_path = Path(prompt_text("Config save path", default=str(Path(output_folder) / "catalogue_index_config.yaml")))
        save_config(config, save_path)
        print(f"Saved config: {save_path}")
    return config


# =============================================================================
# Main extraction process
# =============================================================================


def validate_config(config: Dict[str, Any]) -> None:
    for key in ["input_pdf", "output_folder", "index_pdf_pages", "required_columns", "sku_detection"]:
        if key not in config:
            raise SystemExit(f"Config missing required key: {key}")
    if "sku" not in config["required_columns"] or "page" not in config["required_columns"]:
        raise SystemExit("Config required_columns must include sku and page.")
    if len(config.get("sku_detection", {}).get("positive_examples", []) or []) < 3:
        raise SystemExit("At least 3 positive SKU examples are required.")
    if "advanced" not in config or "data_bottom" not in config["advanced"]:
        raise SystemExit("Config advanced.data_bottom is required. This is the manual bottom cutoff for the index table.")
    if has_center_aligned_columns(config) and config.get("extraction_mode") != "manual":
        log("Centre-aligned column detected. Manual coordinate mode is required for safe extraction.")
        prompt_manual_blocks(config)


def make_source_catalogue_page(pdf_page: int, config: Dict[str, Any]) -> str:
    offset = config.get("catalogue_page_offset")
    if offset in (None, ""):
        return ""
    try:
        return str(int(pdf_page) + int(offset))
    except Exception:
        return ""


def run_extraction(config: Dict[str, Any]) -> Dict[str, Any]:
    validate_config(config)
    input_pdf = Path(config["input_pdf"])
    output_folder = Path(config["output_folder"])
    output_folder.mkdir(parents=True, exist_ok=True)
    pages = parse_page_range(config["index_pdf_pages"])
    sku_rx = compile_regex(config.get("sku_detection", {}).get("sku_regex", ""), "SKU")
    page_rx = compile_regex(config.get("page_detection", {}).get("page_regex", ""), "page")
    advanced = config.get("advanced", {}) or {}

    index_rows: List[Dict[str, Any]] = []
    unresolved_rows: List[Dict[str, Any]] = []
    page_diagnostics: List[Dict[str, Any]] = []
    header_detection: List[Dict[str, Any]] = []
    raw_sku_audit: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    issue_pages = set()
    debug_blocks: Dict[int, List[TableBlock]] = defaultdict(list)

    log(f"Opening PDF: {input_pdf}")
    with pdfplumber.open(str(input_pdf)) as pdf:
        total = len(pdf.pages)
        log(f"PDF has {total} pages. Processing {len(pages)} configured index pages.")
        for idx, pdf_page in enumerate(pages, start=1):
            start = time.perf_counter()
            required_issues: List[str] = []
            optional_warnings: List[str] = []
            raw_audit_issues: List[str] = []
            page_rows: List[Dict[str, Any]] = []
            page_raw_audit: List[Dict[str, Any]] = []
            log(f"Page {idx}/{len(pages)}: PDF page {pdf_page} starting")
            if pdf_page < 1 or pdf_page > total:
                errors.append({"source_pdf_page": pdf_page, "error": "configured_pdf_page_out_of_range"})
                issue_pages.add(pdf_page)
                continue
            page = pdf.pages[pdf_page - 1]
            try:
                words = page.extract_words(
                    x_tolerance=advanced.get("x_tolerance", 1),
                    y_tolerance=advanced.get("y_tolerance", 3),
                    keep_blank_chars=False,
                    use_text_flow=False,
                )
                if config.get("extraction_mode", "auto") == "manual":
                    blocks = manual_blocks_from_config(float(page.width), float(page.height), pdf_page, config)
                    header_records: List[Dict[str, Any]] = []
                else:
                    blocks, header_records = detect_auto_blocks(float(page.width), pdf_page, words, float(page.height), config)
                    header_detection.extend(header_records)
                debug_blocks[pdf_page].extend(blocks)
                expected_blocks = int(config.get("expected_table_blocks_per_page", 1))
                if len(blocks) != expected_blocks:
                    required_issues.append(f"expected_blocks={expected_blocks};detected_blocks={len(blocks)}")
                for block in blocks:
                    rows, block_required, block_optional, block_raw = extract_rows_from_block(pdf_page, words, block, config, sku_rx, page_rx)
                    page_rows.extend(rows)
                    index_rows.extend(rows)
                    raw_sku_audit.extend(block_raw)
                    page_raw_audit.extend(block_raw)
                    required_issues.extend(block_required)
                    optional_warnings.extend(block_optional)
            except Exception as exc:
                errors.append({"source_pdf_page": pdf_page, "error": repr(exc)})
                issue_pages.add(pdf_page)
                continue

            # Raw audit issues are page-level, based on SKU column only.
            structured_norms = {r.get("sku_normalized", "") for r in page_rows if r.get("sku_normalized") and r.get("catalogue_page_original")}
            raw_norms = {r.get("sku_normalized", "") for r in page_raw_audit if r.get("sku_normalized")}
            missing_raw = sorted(raw_norms - structured_norms)
            if missing_raw:
                raw_audit_issues.append(f"sku_column_raw_candidates_not_in_structured:{len(missing_raw)}")
            for row in page_rows:
                if row.get("confidence_status") == "needs_review":
                    unresolved_rows.append(row)
            if required_issues or raw_audit_issues:
                issue_pages.add(pdf_page)
            elapsed = time.perf_counter() - start
            page_diagnostics.append({
                "source_pdf_page": pdf_page,
                "source_catalogue_page": make_source_catalogue_page(pdf_page, config),
                "status": "needs_review" if required_issues or raw_audit_issues else "ok",
                "expected_table_blocks": int(config.get("expected_table_blocks_per_page", 1)),
                "detected_table_blocks": len(blocks),
                "structured_rows": len(page_rows),
                "structured_confirmed_rows": sum(1 for r in page_rows if r.get("confidence_status") == "confirmed"),
                "sku_column_raw_audit_rows": len(page_raw_audit),
                "sku_column_raw_audit_unaccounted": len(missing_raw),
                "required_issue_count": len(dedupe_preserve(required_issues)),
                "optional_warning_count": len(dedupe_preserve(optional_warnings)),
                "raw_audit_issue_count": len(raw_audit_issues),
                "required_issues": multiline(required_issues),
                "optional_warnings": multiline(optional_warnings),
                "raw_audit_issues": multiline(raw_audit_issues),
                "debug_image_path": f"debug_images/debug_pdf_page_{pdf_page}.png" if (required_issues or raw_audit_issues) else "",
                "elapsed_seconds": round(elapsed, 3),
            })
            log(f"Page {idx}/{len(pages)}: PDF page {pdf_page} done - blocks={len(blocks)}, rows={len(page_rows)}, required_issues={len(required_issues)}, raw_issues={len(raw_audit_issues)}, {elapsed:.2f}s")

        raw_added_rows = add_raw_column_candidates(index_rows, raw_sku_audit, config)
        unresolved_rows.extend(raw_added_rows)
        for r in raw_added_rows:
            try:
                issue_pages.add(int(r.get("source_pdf_page", 0)))
            except Exception:
                pass

    sku_registry = build_sku_registry(index_rows, config)
    validation_groups = validate_groups(index_rows, config)
    for result in validation_groups:
        if result.get("status") != "found":
            if result.get("matched_source_pdf_page"):
                issue_pages.add(int(result["matched_source_pdf_page"]))

    outputs = {
        "sku_registry": sku_registry,
        "index_rows": index_rows,
        "unresolved_rows": unresolved_rows,
        "page_diagnostics": page_diagnostics,
        "raw_sku_column_audit": raw_sku_audit,
        "header_detection": header_detection,
        "validation_groups": validation_groups,
        "extractor_errors": errors,
        "run_summary": build_run_summary(sku_registry, index_rows, unresolved_rows, page_diagnostics, raw_sku_audit, errors),
    }
    write_outputs(outputs, output_folder, config)
    if config.get("debug_images", {}).get("enabled", True):
        generate_debug_images(input_pdf, output_folder, sorted(issue_pages), debug_blocks, config)
    log("Extraction complete.")
    log(f"Main output: {output_folder / 'sku_registry.csv'}")
    return outputs


# =============================================================================
# Outputs
# =============================================================================


def build_run_summary(sku_registry, index_rows, unresolved_rows, page_diagnostics, raw_audit, errors) -> List[Dict[str, Any]]:
    return [
        {"metric": "sku_registry_rows", "value": len(sku_registry)},
        {"metric": "index_rows", "value": len(index_rows)},
        {"metric": "unresolved_rows", "value": len(unresolved_rows)},
        {"metric": "pages_processed", "value": len(page_diagnostics)},
        {"metric": "pages_needing_review", "value": sum(1 for r in page_diagnostics if r.get("status") == "needs_review")},
        {"metric": "sku_column_raw_audit_rows", "value": len(raw_audit)},
        {"metric": "raw_audit_candidates_added", "value": sum(1 for r in raw_audit if r.get("added_to_registry_from_raw_text"))},
        {"metric": "extractor_errors", "value": len(errors)},
    ]


def fieldnames_for_rows(rows: List[Dict[str, Any]], preferred: Optional[List[str]] = None) -> List[str]:
    fields: List[str] = []
    seen = set()
    for f in preferred or []:
        if f not in seen:
            seen.add(f)
            fields.append(f)
    for row in rows:
        for f in row.keys():
            if f not in seen:
                seen.add(f)
                fields.append(f)
    return fields


def write_csv(path: Path, rows: List[Dict[str, Any]], preferred: Optional[List[str]] = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = fieldnames_for_rows(rows, preferred)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_review_workbook(path: Path, outputs: Dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        log(f"Could not create review workbook because openpyxl is unavailable: {exc}")
        return

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    sheet_map = [
        ("Run Summary", outputs.get("run_summary", [])),
        ("Page Diagnostics", outputs.get("page_diagnostics", [])),
        ("Index Rows Review", outputs.get("index_rows", [])),
        ("Unresolved Rows", outputs.get("unresolved_rows", [])),
        ("Raw SKU Column Audit", outputs.get("raw_sku_column_audit", [])),
        ("Header Detection", outputs.get("header_detection", [])),
        ("Validation Groups", outputs.get("validation_groups", [])),
        ("Extractor Errors", outputs.get("extractor_errors", [])),
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet_name, rows in sheet_map:
        ws = wb.create_sheet(sheet_name[:31])
        rows = rows or []
        fields = fieldnames_for_rows(rows)
        if not fields:
            fields = ["message"]
            rows = [{"message": "No rows"}]
        ws.append(fields)
        for row in rows:
            ws.append([safe_excel_text(row.get(f, "")) for f in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)
        for col_idx, col in enumerate(fields, start=1):
            letter = get_column_letter(col_idx)
            max_len = len(str(col))
            for cell in ws[letter][:200]:
                try:
                    max_len = max(max_len, min(60, len(str(cell.value or ""))))
                except Exception:
                    pass
            width = min(45, max(10, max_len + 2))
            if any(key in col.lower() for key in ["issues", "warnings", "reason", "text", "raw", "source"]):
                width = min(55, max(width, 30))
            ws.column_dimensions[letter].width = width
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            ws.row_dimensions[row_idx].height = 35
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_outputs(outputs: Dict[str, Any], output_folder: Path, config: Dict[str, Any]) -> None:
    write_csv(output_folder / "sku_registry.csv", outputs["sku_registry"], preferred=[
        "sku", "sku_raw_examples", "sku_normalized", "catalogue_pages_original", "catalogue_pages_normalized",
        "source_pdf_pages", "source_catalogue_pages", "source_table_blocks", "source_methods",
        "confidence_status", "required_issues", "optional_warnings", "occurrence_count",
    ])
    write_csv(output_folder / "index_rows.csv", outputs["index_rows"], preferred=[
        "sku_raw", "sku", "sku_normalized", "catalogue_page_original", "catalogue_page_normalized",
        "source_pdf_page", "source_catalogue_page", "source_table_block", "row_number_in_block",
        "confidence_status", "required_issues", "optional_warnings", "source_row_text",
    ])
    write_review_workbook(output_folder / "extraction_review_workbook.xlsx", outputs)
    if bool(config.get("review_files", {}).get("write_separate_csvs", False)):
        for name, rows in outputs.items():
            if name not in {"sku_registry", "index_rows"}:
                write_csv(output_folder / f"{name}.csv", rows)


# =============================================================================
# Debug images
# =============================================================================


def int_box(box: Tuple[float, float, float, float]) -> Tuple[int, int, int, int]:
    x0, y0, x1, y1 = box
    return (int(round(x0)), int(round(y0)), int(round(x1)), int(round(y1)))


def generate_debug_images(input_pdf: Path, output_folder: Path, issue_pages: List[int], debug_blocks: Dict[int, List[TableBlock]], config: Dict[str, Any]) -> None:
    if not issue_pages:
        log("No issue pages found, so no debug images generated.")
        return
    debug_folder = output_folder / "debug_images"
    debug_folder.mkdir(parents=True, exist_ok=True)
    log(f"Generating debug images for {len(issue_pages)} issue page(s).")
    try:
        with pdfplumber.open(str(input_pdf)) as pdf:
            for pdf_page in issue_pages:
                if pdf_page < 1 or pdf_page > len(pdf.pages):
                    continue
                page = pdf.pages[pdf_page - 1]
                try:
                    im = page.to_image(resolution=120)
                    for block in debug_blocks.get(pdf_page, []):
                        im.draw_rect(int_box((block.block_x0, block.data_top, block.block_x1, block.data_bottom)), stroke="red", stroke_width=2)
                        for field, zone in block.column_zones.items():
                            im.draw_rect(int_box((zone.x0, block.data_top, zone.x1, block.data_bottom)), stroke="blue", stroke_width=1)
                    out_path = debug_folder / f"debug_pdf_page_{pdf_page}.png"
                    im.save(str(out_path), format="PNG")
                    # Add labels after pdfplumber saves if PIL is available.
                    if ImageDraw is not None and config.get("debug_images", {}).get("label_zones", True):
                        from PIL import Image
                        img = Image.open(out_path)
                        draw = ImageDraw.Draw(img)
                        scale_x = img.width / float(page.width)
                        scale_y = img.height / float(page.height)
                        for block in debug_blocks.get(pdf_page, []):
                            for field, zone in block.column_zones.items():
                                draw.text((int(zone.x0 * scale_x) + 2, int(block.data_top * scale_y) + 2), f"B{block.block_number} {field}", fill=(0, 0, 255))
                        img.save(out_path)
                except Exception as exc:
                    log(f"Could not generate debug image for PDF page {pdf_page}: {exc}")
    except Exception as exc:
        log(f"Could not open PDF for debug image generation: {exc}")


# =============================================================================
# CLI
# =============================================================================


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract SKU/page registry from selectable-text PDF catalogue index pages.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""
        Examples
        --------
        Interactive setup:
          py catalogue_index_extractor.py --interactive

        Run from saved config:
          py catalogue_index_extractor.py --config catalogue_index_config.yaml
        """),
    )
    parser.add_argument("--config", type=str, help="Path to YAML or JSON config file.")
    parser.add_argument("--interactive", action="store_true", help="Prompt for settings interactively and optionally save config.")
    parser.add_argument("--input", type=str, help="Override input_pdf from config.")
    parser.add_argument("--output", type=str, help="Override output_folder from config.")
    parser.add_argument("--pages", type=str, help="Override index_pdf_pages from config.")
    parser.add_argument("--manual-coordinates", action="store_true", help="Prompt for manual coordinate blocks before running.")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> None:
    args = parse_args(argv)
    if args.interactive or not args.config:
        config = build_interactive_config()
    else:
        config = load_config(Path(args.config))
    if args.input:
        config["input_pdf"] = args.input
    if args.output:
        config["output_folder"] = args.output
    if args.pages:
        config["index_pdf_pages"] = args.pages
    if args.manual_coordinates:
        prompt_manual_blocks(config)
    output_folder = Path(config["output_folder"])
    output_folder.mkdir(parents=True, exist_ok=True)
    save_config(config, output_folder / "catalogue_index_config_used.yaml")
    run_extraction(config)


if __name__ == "__main__":
    main()
